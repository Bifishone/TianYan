#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TianYan - 天眼查企业信息采集工具
从天眼查获取公司ICP备案、APP、微信公众号、微博信息
"""

import argparse
import requests
import time
import sys
import os
import re as _re
import urllib.parse
import json
import pandas as pd
from bs4 import BeautifulSoup
from colorama import Fore, Style, init
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import configparser

# ── 版本与作者信息 ─────────────────────────────────────────────
__version__ = "1.0"
__author__ = "一只鱼"
__github__ = "https://github.com/Bifishone/"

# ── 常量 ───────────────────────────────────────────────────────
DEFAULT_TYCID = "e431a7203ba411f09e9761d98b79f5e0"
REQUEST_DELAY = 1       # 分页请求间隔（秒）
COMPANY_DELAY = 2       # 公司之间请求间隔（秒）
PAGE_SIZE = 100         # 每页数据量
REQUEST_TIMEOUT = 15    # 请求超时（秒）

# 初始化colorama以支持Windows终端颜色
init(autoreset=True)


# ══════════════════════════════════════════════════════════════════
#  Banner 显示
# ══════════════════════════════════════════════════════════════════

def show_banner():
    """显示程序横幅和作者信息"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    banner_path = os.path.join(current_dir, "banner.txt")

    banner_text = ""
    if os.path.isfile(banner_path):
        try:
            with open(banner_path, "r", encoding="utf-8") as f:
                banner_text = f.read()
        except Exception:
            pass

    if banner_text.strip():
        # 逐行染色：上半部分用青色，中间高亮，底部用蓝色
        lines = banner_text.strip("\n").split("\n")
        total = len(lines)
        for idx, line in enumerate(lines):
            if idx < total * 0.4:
                color = Fore.CYAN
            elif idx < total * 0.8:
                color = Fore.WHITE + Style.BRIGHT
            else:
                color = Fore.BLUE
            print(f"{color}{line}{Style.RESET_ALL}")
    else:
        # 后备：纯文本 Banner
        print(f"{Fore.CYAN}{Style.BRIGHT}")
        print(r"  _____ _             __   __")
        print(r" |_   _(_)_ __  __ _  \ \ / /_ _ _ __")
        print(r"   | | | | '_ \/ _` |  \ V / _` | '_ \ ")
        print(r"   | | | | | | | (_| |  | | (_| | | | |")
        print(r"   |_| |_|_| |_|\__,_|  |_|\__,_|_| |_|")
        print(f"{Style.RESET_ALL}")

    # 作者 & 版本信息
    print(f"{Fore.YELLOW}{'─' * 56}{Style.RESET_ALL}")
    print(f"  {Fore.WHITE}TianYan  v{__version__}{Style.RESET_ALL}")
    print(f"  {Fore.WHITE}Author : {__author__}{Style.RESET_ALL}")
    print(f"  {Fore.WHITE}GitHub : {__github__}{Style.RESET_ALL}")
    print(f"{Fore.YELLOW}{'─' * 56}{Style.RESET_ALL}")
    print()


# ══════════════════════════════════════════════════════════════════
#  HTTP 请求头构建（消除重复代码）
# ══════════════════════════════════════════════════════════════════

def build_api_headers():
    """构建天眼查 API 请求头（capi.tianyancha.com）"""
    return {
        "Host": "capi.tianyancha.com",
        "X-Auth-Token": X_AUTH_TOKEN,
        "Sec-Ch-Ua-Platform": "\"Windows\"",
        "Sec-Ch-Ua": "\"Google Chrome\";v=\"141\", \"Not?A_Brand\";v=\"8\", \"Chromium\";v=\"141\"",
        "Sec-Ch-Ua-Mobile": "?0",
        "X-Tycid": TYCID,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json",
        "Version": "TYC-Web",
        "Origin": "https://www.tianyancha.com",
        "Sec-Fetch-Site": "same-site",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Referer": "https://www.tianyancha.com/",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Priority": "u=1, i",
        "Connection": "keep-alive",
    }


def build_web_headers():
    """构建天眼查网页请求头（www.tianyancha.com）"""
    return {
        "Host": "www.tianyancha.com",
        "Cookie": COOKIES,
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
    }


# ══════════════════════════════════════════════════════════════════
#  配置加载
# ══════════════════════════════════════════════════════════════════

def load_config():
    """加载配置文件中的cookie、x-auth-token和tycid（支持含特殊字符的配置值）"""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(current_dir, "config.ini")

    config = configparser.ConfigParser(interpolation=None)
    try:
        config.read(config_path, encoding="utf-8")

        if "tianyancha" not in config.sections():
            print(f"{Fore.RED}配置文件错误：未找到[tianyancha]部分，请检查{config_path}格式{Style.RESET_ALL}")
            sys.exit(1)

        cookies = config.get("tianyancha", "cookies", fallback="", raw=True)
        x_auth_token = config.get("tianyancha", "x-auth-token", fallback="", raw=True)
        tycid = config.get("tianyancha", "tycid", fallback="", raw=True)

        if not cookies.strip():
            print(f"{Fore.RED}配置错误：cookies不存在或为空，请检查{config_path}{Style.RESET_ALL}")
            sys.exit(1)

        if not x_auth_token.strip():
            print(f"{Fore.RED}配置错误：x-auth-token不存在或为空，请检查{config_path}{Style.RESET_ALL}")
            sys.exit(1)

        if not tycid.strip():
            match = _re.search(r"TYCID=([a-f0-9]+)", cookies)
            if match:
                tycid = match.group(1)
                print(f"{Fore.CYAN}从cookies中自动提取TYCID: {tycid}{Style.RESET_ALL}")
            else:
                tycid = DEFAULT_TYCID
                print(f"{Fore.YELLOW}提示：未配置tycid且无法从cookies中提取，使用默认值。建议在config.ini的[tianyancha]部分添加tycid配置项{Style.RESET_ALL}")

        return cookies, x_auth_token, tycid

    except FileNotFoundError:
        print(f"{Fore.RED}配置文件错误：未找到{config_path}文件，请确保该文件存在{Style.RESET_ALL}")
        sys.exit(1)
    except configparser.ParsingError as e:
        print(f"{Fore.RED}配置文件解析错误：{config_path}可能包含特殊字符（如%），请检查格式。错误详情：{e}{Style.RESET_ALL}")
        sys.exit(1)
    except Exception as e:
        print(f"{Fore.RED}加载配置文件{config_path}时发生错误：{e}{Style.RESET_ALL}")
        sys.exit(1)


# 加载全局配置
COOKIES, X_AUTH_TOKEN, TYCID = load_config()


# ══════════════════════════════════════════════════════════════════
#  公司搜索
# ══════════════════════════════════════════════════════════════════

def search_company(company_name):
    """搜索公司并返回其ID（优先使用API，失败后尝试HTML解析）"""
    # ── 方法1: API 搜索 ──
    try:
        timestamp = int(time.time() * 1000)
        encoded_name = urllib.parse.quote(company_name)
        search_url = (
            f"https://capi.tianyancha.com/cloud-tempest/search/suggest/v3"
            f"?_={timestamp}&keyword={encoded_name}"
        )
        response = requests.get(search_url, headers=build_api_headers(), timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        data = response.json()

        company_list = (
            data.get("data", {}).get("companyList")
            or data.get("data", {}).get("items")
            or data.get("data", {}).get("resultList")
            or []
        )
        if isinstance(company_list, list) and company_list:
            for item in company_list:
                if isinstance(item, dict):
                    company_id = str(
                        item.get("id", "") or item.get("graphId", "") or item.get("companyId", "")
                    )
                    if company_id:
                        return company_id
            first_id = str(company_list[0]) if isinstance(company_list[0], (int, str)) else None
            if first_id and first_id.isdigit():
                return first_id
    except Exception:
        pass  # 静默回退到 HTML 解析

    # ── 方法2: HTML 页面解析 ──
    try:
        encoded_name = urllib.parse.quote(company_name)
        url = f"https://www.tianyancha.com/search?key={encoded_name}"
        response = requests.get(url, headers=build_web_headers(), timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        selectors = [
            "a.index_alink__zcia5.link-click",
            "a[href*=\"/company/\"]",
            "div.search-result-single a",
            "div.result-list a[href*=\"/company/\"]",
            "a.name",
            "a[class*=\"alink\"]",
        ]

        for selector in selectors:
            for a_tag in soup.select(selector):
                href = a_tag.get("href", "")
                if "/company/" in href:
                    try:
                        parts = href.split("/company/")[-1].split("/")
                        for part in parts:
                            if part and part.replace("-", "").replace("_", "").isalnum() and len(part) > 3:
                                return part
                    except Exception:
                        continue

        # 最后尝试：所有含 /company/ 的链接
        for a_tag in soup.find_all("a", href=True):
            href = a_tag["href"]
            if "/company/" in href:
                parts = href.split("/company/")[-1].split("/")
                for part in parts:
                    if part and len(part) > 3 and not part.startswith("#"):
                        return part
        return None

    except Exception as e:
        print(f"{Fore.RED}搜索 {company_name} 时出错: {e}{Style.RESET_ALL}")
        return None


# ══════════════════════════════════════════════════════════════════
#  通用分页数据获取器
# ══════════════════════════════════════════════════════════════════

def _fetch_paginated_data(url_template, label, data_key_path, item_processor,
                           extra_params=None, timeout=REQUEST_TIMEOUT):
    """
    通用分页数据获取器。

    参数:
        url_template: URL 模板，使用 {timestamp} 和 {page_num} 占位
        label: 数据名称（用于日志打印）
        data_key_path: 从 response.json() 中提取列表的 JSON 路径，如 ["data", "items"]
        item_processor: 处理单个 item 的回调，返回 dict 或 None
        extra_params: 额外 URL 参数字典
        timeout: 请求超时

    返回:
        list[dict]: 所有页收集到的数据行
    """
    collected = []
    page_num = 1
    has_more = True

    print(f"  开始获取{label}数据...")

    while has_more:
        timestamp = int(time.time() * 1000)
        url = url_template.format(timestamp=timestamp, page_num=page_num)
        if extra_params:
            url += "&" + urllib.parse.urlencode(extra_params)

        try:
            response = requests.get(url, headers=build_api_headers(), timeout=timeout)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"{Fore.RED}  第{page_num}页{label}请求失败: {e}，停止获取{Style.RESET_ALL}")
            break

        try:
            data = response.json()
            if not isinstance(data, dict):
                print(f"{Fore.YELLOW}  第{page_num}页{label}响应非字典格式，停止{Style.RESET_ALL}")
                break
        except json.JSONDecodeError:
            print(f"{Fore.YELLOW}  第{page_num}页{label}响应非有效JSON，停止{Style.RESET_ALL}")
            break

        # 按路径提取数据列表
        items = data
        for key in data_key_path:
            items = items.get(key, {}) if isinstance(items, dict) else []
        if not isinstance(items, list):
            items = []

        if items:
            print(f"  已获取第{page_num}页{label}数据，共{len(items)}条")
            for item in items:
                if isinstance(item, dict):
                    row = item_processor(item)
                    if row:
                        collected.append(row)
            page_num += 1
            time.sleep(REQUEST_DELAY)
        else:
            has_more = False
            if page_num == 1:
                print(f"  未找到{label}数据")
            else:
                print(f"  {label}数据获取完成，共{page_num - 1}页")

    return collected


# ══════════════════════════════════════════════════════════════════
#  ICP 备案信息
# ══════════════════════════════════════════════════════════════════

def get_company_info(company_id, company_name):
    """获取ICP备案信息"""
    domains = []

    def process_item(item):
        web_site = (item.get("webSite") or [None])[0] if item.get("webSite") else ""
        domain = item.get("ym", "")
        if domain:
            domains.append(domain)
        return {
            "公司名称": company_name,
            "公司ID": company_id,
            "审核日期": item.get("examineDate", ""),
            "网站名称": item.get("webName", ""),
            "网站首页": web_site,
            "域名": domain,
            "网站备案/许可证号": item.get("liscense", ""),
        }

    url_template = (
        "https://capi.tianyancha.com/cloud-intellectual-property/intellectualProperty/icpRecordList"
        "?_={timestamp}&id={company_id}&pageSize={page_size}&pageNum={page_num}"
    ).replace("{company_id}", str(company_id)).replace("{page_size}", str(PAGE_SIZE))

    extracted_data = _fetch_paginated_data(
        url_template=url_template,
        label="ICP备案",
        data_key_path=["data", "item"],
        item_processor=process_item,
    )

    return extracted_data, domains


# ══════════════════════════════════════════════════════════════════
#  APP 信息
# ══════════════════════════════════════════════════════════════════

def get_app_info(company_id, company_name):
    """获取APP信息"""

    def process_item(item):
        return {
            "公司名称": company_name,
            "产品名称": item.get("name", ""),
            "产品简称": item.get("filterName", ""),
            "产品分类": item.get("type", ""),
            "领域": item.get("classes", ""),
            "图标链接": item.get("icon", ""),
        }

    url_template = (
        "https://capi.tianyancha.com/cloud-business-state/v3/ar/appbkinfo"
        "?_={timestamp}&id={company_id}&pageSize={page_size}&pageNum={page_num}"
    ).replace("{company_id}", str(company_id)).replace("{page_size}", str(PAGE_SIZE))

    return _fetch_paginated_data(
        url_template=url_template,
        label="APP",
        data_key_path=["data", "items"],
        item_processor=process_item,
    )


# ══════════════════════════════════════════════════════════════════
#  微信公众号信息
# ══════════════════════════════════════════════════════════════════

def get_wechat_info(company_id, company_name):
    """获取微信公众号信息"""

    def process_item(item):
        return {
            "公司名称": company_name,
            "微信公众号": item.get("title", ""),
            "微信号": item.get("publicNum", ""),
            "简介": item.get("recommend", ""),
            "二维码链接": item.get("codeImg", ""),
            "图标链接": item.get("titleImgURL", ""),
        }

    url_template = (
        "https://capi.tianyancha.com/cloud-business-state/wechat/list"
        "?_={timestamp}&graphId={company_id}&pageSize={page_size}&pageNum={page_num}"
    ).replace("{company_id}", str(company_id)).replace("{page_size}", str(PAGE_SIZE))

    return _fetch_paginated_data(
        url_template=url_template,
        label="微信公众号",
        data_key_path=["data", "resultList"],
        item_processor=process_item,
    )


# ══════════════════════════════════════════════════════════════════
#  微博信息
# ══════════════════════════════════════════════════════════════════

def get_weibo_info(company_id, company_name):
    """获取微博信息"""

    def process_item(item):
        tags = item.get("tags", [])
        return {
            "公司名称": company_name,
            "微博名称": item.get("name", ""),
            "行业类别": ",".join(tags) if isinstance(tags, list) else (tags or ""),
            "简介": item.get("info", ""),
            "微博链接": item.get("href", ""),
            "图标链接": item.get("ico", ""),
        }

    url_template = (
        "https://capi.tianyancha.com/cloud-business-state/weibo/list"
        "?_={timestamp}&graphId={company_id}&pageSize={page_size}&pageNum={page_num}"
    ).replace("{company_id}", str(company_id)).replace("{page_size}", str(PAGE_SIZE))

    return _fetch_paginated_data(
        url_template=url_template,
        label="微博",
        data_key_path=["data", "result"],
        item_processor=process_item,
    )


# ══════════════════════════════════════════════════════════════════
#  文件读写
# ══════════════════════════════════════════════════════════════════

def read_company_names(file_path):
    """读取公司名称列表"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return [line.strip() for line in f if line.strip()]
    except Exception as e:
        print(f"{Fore.RED}读取文件 {file_path} 时出错: {e}{Style.RESET_ALL}")
        sys.exit(1)


def write_results(all_data, domains, app_data, wechat_data, weibo_data, xlsx_file, domains_file):
    """写入Excel和域名文件"""
    # 确保输出目录存在
    for path in (xlsx_file, domains_file):
        if path:
            out_dir = os.path.dirname(path)
            if out_dir:
                os.makedirs(out_dir, exist_ok=True)

    # ── 写入 Excel ──
    if (all_data or app_data or wechat_data or weibo_data) and xlsx_file:
        try:
            with pd.ExcelWriter(xlsx_file, engine="openpyxl") as writer:
                if all_data:
                    df_icp = pd.DataFrame(all_data).drop(columns=["公司ID"]).fillna("—")
                    df_icp.to_excel(writer, index=False, sheet_name="ICP备案数据")
                    format_worksheet(writer.sheets["ICP备案数据"], "ICP备案数据")

                if app_data:
                    df_app = pd.DataFrame(app_data).fillna("—")
                    df_app = df_app[["公司名称", "产品名称", "产品简称", "产品分类", "领域", "图标链接"]]
                    df_app.to_excel(writer, index=False, sheet_name="APP信息")
                    format_worksheet(writer.sheets["APP信息"], "APP信息")

                if wechat_data:
                    df_wechat = pd.DataFrame(wechat_data).fillna("—")
                    df_wechat = df_wechat[["公司名称", "微信公众号", "微信号", "简介", "二维码链接", "图标链接"]]
                    df_wechat.to_excel(writer, index=False, sheet_name="微信公众号信息")
                    format_worksheet(writer.sheets["微信公众号信息"], "微信公众号信息")

                if weibo_data:
                    df_weibo = pd.DataFrame(weibo_data).fillna("—")
                    df_weibo = df_weibo[["公司名称", "微博名称", "行业类别", "简介", "微博链接", "图标链接"]]
                    df_weibo.to_excel(writer, index=False, sheet_name="微博信息")
                    format_worksheet(writer.sheets["微博信息"], "微博信息")

            print(f"{Fore.GREEN}数据已成功导出到 {xlsx_file}{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}写入Excel文件 {xlsx_file} 时出错: {e}{Style.RESET_ALL}")

    # ── 写入域名文件 ──
    if domains and domains_file:
        try:
            with open(domains_file, "w", encoding="utf-8") as f:
                for domain in domains:
                    f.write(domain + "\n")
            print(f"{Fore.GREEN}域名已成功导出到 {domains_file}{Style.RESET_ALL}")
        except Exception as e:
            print(f"{Fore.RED}写入域名文件 {domains_file} 时出错: {e}{Style.RESET_ALL}")


# ══════════════════════════════════════════════════════════════════
#  Excel 格式化
# ══════════════════════════════════════════════════════════════════

def format_worksheet(worksheet, sheet_name):
    """格式化Excel工作表，根据不同工作表使用差异化配色"""
    sheet_colors = {
        "ICP备案数据": "4F81BD",
        "APP信息": "548235",
        "微信公众号信息": "7030A0",
        "微博信息": "C00000",
    }
    theme_color = sheet_colors.get(sheet_name, "4F81BD")

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    content_font = Font(name="微软雅黑", size=10, color="333333")
    content_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    empty_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    empty_font = Font(name="微软雅黑", size=10, color="999999")

    key_col_names = {"网站备案/许可证号", "域名", "微博链接", "图标链接", "二维码链接"}

    # 表头
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 内容行
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        row_fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill

        for cell in row:
            cell.border = thin_border
            cell.alignment = content_alignment

            if not cell.value or str(cell.value).strip() in ("", "—"):
                cell.fill = empty_fill
                cell.font = empty_font
            else:
                cell.fill = row_fill
                cell.font = content_font

                col_name = worksheet.cell(row=1, column=cell.column).value
                if col_name in key_col_names:
                    cell.font = Font(name="微软雅黑", size=10, color=theme_color)

    # 自动列宽
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        header_value = str(worksheet.cell(row=1, column=col[0].column).value)
        max_length = max(
            len(header_value),
            max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0),
        )
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


# ══════════════════════════════════════════════════════════════════
#  主入口
# ══════════════════════════════════════════════════════════════════

def main():
    """主函数"""
    # ── 显示 Banner ──
    show_banner()

    parser = argparse.ArgumentParser(
        description="TianYan — 从天眼查获取公司ICP备案、APP、微信公众号、微博信息",
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--file", help="包含公司名称的文本文件路径，每行一个公司名称")
    group.add_argument("--CompanyName", help="单个公司名称")
    parser.add_argument("--out-xlsx", help="输出结果的Excel文件路径（含4张工作表）")
    parser.add_argument("--out-domains", help="输出域名列表的文本文件路径")
    parser.add_argument("--include-app", action="store_true", help="是否包含APP信息查询")
    parser.add_argument("--include-wechat", action="store_true", help="是否包含微信公众号信息查询")
    parser.add_argument("--include-weibo", action="store_true", help="是否包含微博信息查询")
    parser.add_argument("--tycid", help="指定X-Tycid请求头（覆盖config.ini中的配置）")

    args = parser.parse_args()

    # 命令行覆盖 tycid
    global TYCID
    if args.tycid:
        TYCID = args.tycid.strip()
        print(f"{Fore.CYAN}使用命令行指定的TYCID: {TYCID[:20]}...{Style.RESET_ALL}")

    if not args.out_xlsx and not args.out_domains:
        print(f"{Fore.RED}错误: 必须指定 --out-xlsx 或 --out-domains 至少一个输出文件{Style.RESET_ALL}")
        sys.exit(1)

    companies = read_company_names(args.file) if args.file else [args.CompanyName]

    all_data, all_domains = [], []
    all_app_data, all_wechat_data, all_weibo_data = [], [], []
    total = len(companies)
    company_ids = {}

    print(f"{Fore.CYAN}开始处理 {total} 家公司...{Style.RESET_ALL}")
    print("-" * 80)

    for i, company in enumerate(companies, 1):
        print(f"{Fore.YELLOW}[{i}/{total}] 处理公司: {company}{Style.RESET_ALL}")

        company_id = search_company(company)
        company_ids[company] = company_id

        if company_id:
            # ICP 备案信息
            print(f"{Fore.BLUE}  处理ICP备案信息...{Style.RESET_ALL}")
            company_data, domains = get_company_info(company_id, company)
            if company_data:
                print(f"{Fore.GREEN}  成功获取 {len(company_data)} 条备案信息{Style.RESET_ALL}")
                all_data.extend(company_data)
                all_domains.extend(domains)
            else:
                print(f"{Fore.YELLOW}  未获取到ICP备案信息{Style.RESET_ALL}")

            # 可选模块
            optional_tasks = [
                (args.include_app, "APP信息", get_app_info, all_app_data),
                (args.include_wechat, "微信公众号信息", get_wechat_info, all_wechat_data),
                (args.include_weibo, "微博信息", get_weibo_info, all_weibo_data),
            ]
            for enabled, label, fetcher, collector in optional_tasks:
                if enabled:
                    print(f"{Fore.BLUE}  处理{label}...{Style.RESET_ALL}")
                    result = fetcher(company_id, company)
                    if result:
                        print(f"{Fore.GREEN}  成功获取 {len(result)} 条{label}{Style.RESET_ALL}")
                        collector.extend(result)
                    else:
                        print(f"{Fore.YELLOW}  未获取到{label}{Style.RESET_ALL}")
        else:
            print(f"{Fore.RED}  未找到该公司信息{Style.RESET_ALL}")

        print("-" * 80)
        if i < total:
            time.sleep(COMPANY_DELAY)

    # 输出结果
    if all_data or all_domains or all_app_data or all_wechat_data or all_weibo_data:
        write_results(all_data, all_domains, all_app_data, all_wechat_data, all_weibo_data,
                       args.out_xlsx, args.out_domains)
    else:
        print(f"{Fore.YELLOW}没有获取到任何数据，不生成输出文件{Style.RESET_ALL}")

    found_ids = sum(1 for cid in company_ids.values() if cid)
    print(
        f"\n{Fore.CYAN}处理完成。共找到 {found_ids}/{total} 家公司的信息，"
        f"获取到 {len(all_data)} 条备案信息，提取到 {len(all_domains)} 个域名，"
        f"获取到 {len(all_app_data)} 条APP信息，{len(all_wechat_data)} 条微信公众号信息，"
        f"{len(all_weibo_data)} 条微博信息{Style.RESET_ALL}"
    )


if __name__ == "__main__":
    main()
